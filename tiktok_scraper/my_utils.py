import json
import openpyxl


def getpath_of_value_in_dictionary(nested_dict, value, prepath=()):
    for k, v in nested_dict.items():
        path = prepath + (k,)
        if v == value: # found value
            return path
        elif hasattr(v, 'items'): # v is a dict
            p = getpath_of_value_in_dictionary(v, value, path) # recursive call
            if p is not None:
                return p


def crint(text):
    HEADER = '\033[95m'
    OKBLUE = '\033[94m'
    OKCYAN = '\033[96m'
    OKGREEN = '\033[92m'
    WARNING = '\033[93m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'
    print(f"{BOLD}{UNDERLINE}{OKGREEN}{text}{ENDC}")
    # print(f"{UNDERLINE}{text}{ENDC}")


def lists_to_excel(data, output_file):  # data is list of lists and first element of each list is title of that column
    # Transpose the data
    transposed_data = list(map(list, zip(*data)))

    # Create a new workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Write transposed data to the worksheet
    for row_num, row_data in enumerate(transposed_data, 1):
        for col_num, cell_value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num, value=cell_value)

    # Save the workbook
    wb.save(output_file)


def dics_to_excel(data, output_file):
    # Create a new workbook and select the active worksheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Check if data is not empty
    if not data:
        print("No data to write.")
        return

    # Write the header row
    headers = list(data[0].keys())
    sheet.append(headers)

    # Write the data rows
    for row in data:
        sheet.append(list(row.values()))

    # Save the workbook to the specified file path
    workbook.save(output_file)
